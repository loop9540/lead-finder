"""
Source: Website technology detection.
Scans company websites from 9at data to detect which investor portal
technology platform they use by looking for links, references, and
login URLs in their HTML.

~42% detection rate in testing.
"""

import csv
import re
import subprocess
import time
import openpyxl

PLATFORM_SIGNATURES = {
    "InvestorFlow": ["investorflow.com"],
    "Dynamo": ["dynamosoftware.com"],
    "Intralinks": ["intralinks.com"],
    "eFront / Investment Cafe": ["efrontcloud.com", "efront.com/en"],
    "AllVue / AltaReturn": ["altareturn.com", "allvuesystems.com"],
    "ARKPES": ["arkpes.com"],
    "Sharesecure": ["sharesecurely.com", "altvia.com"],
    "Juniper Square": ["junipersquare.com"],
    "Carta": ["carta.com/investor"],
    "SS&C Fund Services": ["sscfundservices.com", "globeop.com", "ssinc.com"],
}

# Competitor fund administrators to scan from 9at
TARGET_ADMINS = {
    "SS&C TECHNOLOGIES", "APEX GROUP", "CITCO", "ALTER DOMUS",
    "STANDISH MANAGEMENT", "IQ-EQ", "FIS GLOBAL", "MAPLES",
    "TMF GROUP INC", "HEDGESERV", "AZTEC GROUP", "CSC",
    "LANGHAM HALL", "OCORIAN", "OPUS FUND SERVICES",
    "TRIDENT TRUST", "STONE COAST FUND SERVICES",
    "HC GLOBAL FUND SERVICES", "PETRA FUNDS GROUP",
    "JTC GROUP", "WAYSTONE", "VISTRA", "FORMIDIUM",
    "NAV CONSULTING, INC", "ADURO ADVISORS",
}

# Skip Gen II's own clients
SKIP_ADMINS = {"GEN II FUND SERVICES"}


def _fetch_html(url: str) -> str:
    """Fetch page HTML using curl."""
    if not url.startswith("http"):
        url = "https://" + url
    url = url.split("\t")[0].split(" ")[0].strip()

    try:
        result = subprocess.run(
            ["curl", "-s", "-L", "--max-time", "8",
             "-H", "User-Agent: Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36",
             url],
            capture_output=True, text=True, timeout=12,
        )
        return result.stdout if result.returncode == 0 else ""
    except (subprocess.TimeoutExpired, Exception):
        return ""


def _detect_platform(html: str) -> list[str]:
    """Detect platform signatures in HTML content."""
    html_lower = html.lower()
    found = []
    for platform, sigs in PLATFORM_SIGNATURES.items():
        for sig in sigs:
            if sig in html_lower:
                found.append(platform)
                break
    return found


def load_9at_companies(filepath: str, max_companies: int = 5000) -> list[dict]:
    """
    Load unique companies from 9at export that use competitor administrators.
    Returns list of dicts with company, admin, website.
    """
    print(f"  [tech-detect] Loading 9at data...")
    wb = openpyxl.load_workbook(filepath, read_only=False)
    ws = wb.active

    companies = {}
    for r in range(3, min(ws.max_row + 1, 120000)):
        admin = ws.cell(r, 22).value  # Administrator
        company = ws.cell(r, 4).value  # Parent Adviser Name
        website = ws.cell(r, 28).value  # All Websites

        if not admin or not company or not website:
            continue

        admin = str(admin).strip()
        company = str(company).strip()
        website = str(website).strip()

        if admin in SKIP_ADMINS:
            continue
        if admin not in TARGET_ADMINS:
            continue
        if company in companies:
            continue

        companies[company] = {
            "company": company,
            "admin": admin,
            "website": website,
        }

        if len(companies) >= max_companies:
            break

    wb.close()
    print(f"  [tech-detect] Found {len(companies)} unique companies with competitor admins + websites")
    return list(companies.values())


def scan_companies(companies: list[dict], progress_callback=None) -> list[dict]:
    """
    Scan company websites to detect technology platforms.
    Returns leads with detected platform.
    """
    results = []
    total = len(companies)
    detected = 0

    for i, comp in enumerate(companies):
        if i % 50 == 0:
            msg = f"  [tech-detect] Scanning {i}/{total} (detected: {detected})..."
            print(msg)
            if progress_callback:
                progress_callback(msg)

        html = _fetch_html(comp["website"])
        if not html:
            continue

        platforms = _detect_platform(html)
        if platforms:
            detected += 1
            for platform in platforms:
                results.append({
                    "company_name": comp["company"],
                    "domain": comp["website"].split("\t")[0].strip(),
                    "platform": platform,
                    "source": "tech_detect",
                })

        # Small delay to be polite
        if i % 10 == 0:
            time.sleep(0.5)

    print(f"  [tech-detect] Done: {detected}/{total} companies detected ({detected*100//max(total,1)}%)")
    return results


def discover_from_9at(filepath: str, max_companies: int = 5000,
                      progress_callback=None) -> list[dict]:
    """
    Full pipeline: load 9at data, scan websites, return leads.
    """
    companies = load_9at_companies(filepath, max_companies)
    if not companies:
        print("  [tech-detect] No companies to scan")
        return []
    return scan_companies(companies, progress_callback)
