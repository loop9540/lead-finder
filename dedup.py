"""
Deduplication logic for leads.
- Normalizes company names for fuzzy matching
- Deduplicates against existing leads (the 200-company list)
- Deduplicates within new results across sources
"""

import re
import os
import openpyxl


def normalize(name: str) -> str:
    """Normalize a company name for comparison."""
    name = name.strip()
    # Split CamelCase into words (e.g., "CapeCapital" -> "Cape Capital")
    name = re.sub(r'([a-z])([A-Z])', r'\1 \2', name)
    name = name.lower()
    # Remove "the " prefix
    if name.startswith("the "):
        name = name[4:]
    # Remove common suffixes
    for suffix in [
        " llc", " lp", " inc", " corp", " ltd", " co",
        " partners", " capital", " management", " advisors",
        " group", " holdings", " fund", " investment", " investments",
        " properties",
    ]:
        if name.endswith(suffix):
            name = name[: -len(suffix)].strip()
    # Remove punctuation and extra whitespace
    name = re.sub(r"[^a-z0-9\s]", "", name)
    name = re.sub(r"\s+", " ", name).strip()
    return name


def load_existing_leads(filepath: str) -> set[str]:
    """
    Load existing company names from an Excel file for dedup.
    Tries to find a column named 'company', 'company_name', 'name', or 'domain'.
    Returns a set of normalized names.
    """
    if not filepath or not os.path.exists(filepath):
        return set()

    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb.active

    # Find the right column
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    col_idx = None
    for i, h in enumerate(header_row):
        if h and str(h).lower().strip() in ("company", "company_name", "name", "domain"):
            col_idx = i
            break

    if col_idx is None:
        # Default to first column
        col_idx = 0

    existing = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[col_idx]:
            existing.add(normalize(str(row[col_idx])))

    wb.close()
    return existing


def _is_junk(name: str) -> bool:
    """Filter out names that are clearly not company names."""
    low = name.lower().strip()

    # Navigation / UI elements / generic terms
    junk_exact = {
        "about", "home", "resources", "partners", "applications",
        "implementation", "infrastructure", "accounting", "big",
        "info", "private", "cfo", "contact", "subscribe",
        "resource center", "leadership team", "managed services",
        "industry expertise", "industry clouds", "client relationships",
        "capital deployment", "capital formation", "investor services",
        "relationship intelligence", "partnership network",
        "integrated application suite", "multi-cloud architecture",
        "private credit", "private equity", "real assets",
        "company logos", "why investorflow?",
        "client services", "our customers", "u.s. dollar",
        "client lens", "preqin",
    }
    if low in junk_exact:
        return True

    # Platform self-references
    platform_names = {
        "dynamo", "dynamo software", "investorflow", "investorflow ai",
        "intralinks", "efront", "arkpes", "ark pes", "allvue",
        "allvue systems", "altvia", "sharesecure",
    }
    if low in platform_names:
        return True

    # Page titles / breadcrumbs (contain " | " or " - " with platform name)
    if " | " in name or re.search(r" - (InvestorFlow|Dynamo|Altvia|Ark|Intralinks|Allvue)", name):
        return True

    # Case study titles (long descriptive sentences)
    if len(name) > 50:
        return True

    # File names / image alt text artifacts
    if name.endswith((".jpg", ".png", ".svg", ".gif", ".pdf")):
        return True
    if name.lower().endswith(" logo.") or name.lower().endswith(" logo"):
        return True

    # Starts with "Customer Logos" or similar image descriptions
    if low.startswith("customer logos"):
        return True

    # HTML entities
    if "&#" in name or "&amp;" in name:
        return True

    # Single generic words
    if len(low.split()) == 1 and low in {
        "security", "compliance", "analytics", "platform", "services",
        "solutions", "insights", "overview", "features", "pricing",
        "blog", "news", "events", "careers", "login", "dashboard",
    }:
        return True

    return False


def deduplicate(leads: list[dict], existing_file: str = "") -> list[dict]:
    """
    Deduplicate leads:
    1. Filter out junk entries (navigation, page titles, platform names)
    2. Remove duplicates within the list (by normalized company_name)
    3. Remove companies already in the existing leads file
    Returns deduplicated list, preferring entries with more data.
    """
    existing = load_existing_leads(existing_file)
    print(f"  [dedup] Loaded {len(existing)} existing leads for deduplication")

    # Pre-filter junk
    leads = [l for l in leads if not _is_junk(l["company_name"])]

    seen = {}
    for lead in leads:
        key = normalize(lead["company_name"])

        # Skip if in existing list
        if key in existing:
            continue

        # Keep the first occurrence, or prefer one with more fields populated
        if key not in seen:
            seen[key] = lead
        else:
            # Prefer the entry with more non-empty values
            existing_filled = sum(1 for v in seen[key].values() if v)
            new_filled = sum(1 for v in lead.values() if v)
            if new_filled > existing_filled:
                seen[key] = lead

    result = list(seen.values())
    removed = len(leads) - len(result)
    print(f"  [dedup] Removed {removed} duplicates, {len(result)} unique leads remaining")
    return result
