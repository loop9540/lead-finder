"""Contact matching module: match lead company names against 9at export contacts."""

import re
import sqlite3
from collections import defaultdict

import openpyxl

from config import DB_PATH

# Suffixes to strip during normalization
STRIP_SUFFIXES = [
    "llc", "llp", "lp", "inc", "corp", "corporation", "ltd", "limited",
    "management", "advisors", "advisor", "advisory", "advisers", "adviser",
    "partners", "partner", "capital", "group", "holdings", "holding",
    "fund", "funds", "investment", "investments", "investing",
    "services", "service", "associates", "association",
    "consulting", "consultants", "solutions", "enterprises",
    "co", "company", "international", "global", "asset",
    "wealth", "financial", "finance", "strategies", "strategy",
    "equity", "ventures", "venture", "real estate",
]

# Pre-compile a pattern that matches any suffix at word boundaries
_SUFFIX_PATTERN = re.compile(
    r'\b(' + '|'.join(re.escape(s) for s in STRIP_SUFFIXES) + r')\b'
)


def _split_camelcase(name: str) -> str:
    """Split camelCase or concatenated words: lcatterton -> l catterton."""
    # Insert space before uppercase letters that follow lowercase letters
    name = re.sub(r'([a-z])([A-Z])', r'\1 \2', name)
    # Insert space between letter and digit transitions
    name = re.sub(r'([a-zA-Z])(\d)', r'\1 \2', name)
    name = re.sub(r'(\d)([a-zA-Z])', r'\1 \2', name)
    return name


def _generate_variants(name: str) -> list[str]:
    """Generate normalized variants of a name to handle concatenated forms.

    For names like 'lcatterton', try splitting at each position to produce
    'l catterton', 'lc atterton', etc. and normalize each.
    """
    variants = [normalize_company_name(name)]
    # Only do this for single-word names (no spaces in original)
    clean = re.sub(r'[^a-z0-9]', '', name.lower())
    if ' ' not in name.strip() and len(clean) >= 6:
        # Try splitting at each position
        for i in range(1, len(clean)):
            left = clean[:i]
            right = clean[i:]
            variant = f"{left} {right}"
            norm = normalize_company_name(variant)
            if norm and norm != variants[0]:
                variants.append(norm)
    return variants


def normalize_company_name(name: str) -> str:
    """Normalize a company name for matching.

    Steps:
    1. Lowercase
    2. Split camelCase
    3. Remove punctuation
    4. Strip common business suffixes
    5. Collapse whitespace and strip
    """
    if not name:
        return ""
    name = name.lower()
    name = _split_camelcase(name)
    # Remove content in parentheses (e.g., "(now TPG Angelo Gordon)")
    name = re.sub(r'\([^)]*\)', '', name)
    # Remove punctuation (keep alphanumeric and spaces)
    name = re.sub(r'[^a-z0-9\s]', ' ', name)
    # Strip suffixes
    name = _SUFFIX_PATTERN.sub('', name)
    # Collapse whitespace
    name = re.sub(r'\s+', ' ', name).strip()
    return name


def load_9at_contacts(filepath: str) -> dict[str, list[dict]]:
    """Load 9at export contacts into a dict keyed by normalized Parent Adviser Name.

    Returns: {normalized_company_name: [list of contact dicts]}
    Each contact dict has: first_name, last_name, title, email, email_status,
                           phone, linkedin, company_name (original).
    """
    wb = openpyxl.load_workbook(filepath, read_only=False)
    ws = wb.active

    contacts = defaultdict(list)
    # Data starts at row 3 (row 1 empty, row 2 headers)
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
        company_raw = row[3]  # Column D: Parent Adviser Name
        if not company_raw or not str(company_raw).strip():
            continue

        company_raw = str(company_raw).strip()
        normalized = normalize_company_name(company_raw)
        if not normalized:
            continue

        contact = {
            "first_name": str(row[0] or "").strip(),
            "last_name": str(row[1] or "").strip(),
            "title": str(row[2] or "").strip(),
            "email": str(row[5] or "").strip(),
            "email_status": str(row[6] or "").strip(),
            "phone": str(row[7] or "").strip(),
            "linkedin": str(row[8] or "").strip(),
            "company_name": company_raw,
        }
        contacts[normalized].append(contact)

    wb.close()
    return dict(contacts)


def _extract_tokens(name: str) -> set[str]:
    """Extract meaningful tokens from a normalized name."""
    return {t for t in name.split() if len(t) > 1}


def _is_good_match(lead_norm: str, nine_norm: str) -> bool:
    """Determine if a lead name and 9at name are a good match.

    Uses strict rules to avoid false positives:
    1. Exact normalized match
    2. Multi-token leads: ALL lead tokens present in 9at name
    3. Single-token leads (>=5 chars): 9at name must also be that single token
       OR lead token appears as a word-boundary substring in 9at and is very distinctive (>=8 chars)
    """
    if lead_norm == nine_norm:
        return True

    lead_tokens = _extract_tokens(lead_norm)
    nine_tokens = _extract_tokens(nine_norm)

    if not lead_tokens or not nine_tokens:
        return False

    # Multi-token lead: all lead tokens must be in 9at tokens
    if len(lead_tokens) >= 2:
        if lead_tokens.issubset(nine_tokens):
            return True
        # Also try: full lead_norm is a word-boundary substring of nine_norm
        if len(lead_norm) >= 8:
            if re.search(r'(?:^|\s)' + re.escape(lead_norm) + r'(?:\s|$)', nine_norm):
                return True
        return False

    # Single-token lead
    single = next(iter(lead_tokens))

    # For single tokens, only match if the 9at name is also just that token
    # (after normalization both reduce to the same thing)
    if lead_norm == nine_norm:
        return True

    # For very distinctive single tokens (>=8 chars like "blackstone", "brookfield"),
    # allow matching where 9at normalized starts with or equals the token
    if len(single) >= 8 and nine_norm.startswith(single):
        return True

    return False


def match_company_names(
    lead_names: list[str],
    nine_at_names: list[str],
    threshold: float = 0.0,
) -> dict[str, list[str]]:
    """Match lead company names to 9at company names.

    Strategy (in order of priority):
    1. Exact normalized match
    2. Token overlap: all lead tokens appear in 9at name
    3. Substring: lead normalized name is contained in 9at normalized name
    4. Reverse substring: 9at normalized name is contained in lead normalized name

    Returns: {lead_company_name: [list of matched 9at normalized names]}
    """
    # Normalize everything
    lead_normalized = {name: normalize_company_name(name) for name in lead_names}
    nine_at_normalized = {name: normalize_company_name(name) for name in nine_at_names}

    # Build a set of 9at normalized names for fast lookup
    nine_at_norm_set = set(nine_at_normalized.values())
    # Build reverse map: normalized -> original names
    nine_at_norm_to_originals = defaultdict(list)
    for orig, norm in nine_at_normalized.items():
        nine_at_norm_to_originals[norm].append(orig)

    matches = {}

    for lead_orig, lead_norm in lead_normalized.items():
        if not lead_norm or len(lead_norm) < 2:
            continue

        matched_norms = set()

        # Strategy 1: Exact match
        if lead_norm in nine_at_norm_set:
            matched_norms.add(lead_norm)

        # Strategy 2 & 3: Token overlap and substring matching
        if not matched_norms:
            lead_tokens = _extract_tokens(lead_norm)
            if not lead_tokens:
                continue

            for nine_norm in nine_at_norm_set:
                if not nine_norm:
                    continue

                nine_tokens = _extract_tokens(nine_norm)

                # Token overlap: all lead tokens found in 9at name
                if lead_tokens and lead_tokens.issubset(nine_tokens):
                    matched_norms.add(nine_norm)
                    continue

                # Substring: lead name inside 9at name
                if len(lead_norm) >= 4 and lead_norm in nine_norm:
                    matched_norms.add(nine_norm)
                    continue

                # Reverse substring: 9at name inside lead name (only for longer 9at names)
                if len(nine_norm) >= 6 and nine_norm in lead_norm:
                    matched_norms.add(nine_norm)
                    continue

        if matched_norms:
            matches[lead_orig] = list(matched_norms)

    return matches


def match_leads_to_contacts(
    leads_db_path: str = DB_PATH,
    nine_at_path: str = "/Users/idavid/Downloads/9at_export_contacts_2026_03_09.xlsx",
) -> list[dict]:
    """Match leads from SQLite DB against 9at export contacts.

    Returns list of dicts with: lead_id, lead_company, lead_domain, lead_platform,
    lead_source, contacts (list of contact dicts), match_key (normalized name matched).
    """
    # Load leads
    conn = sqlite3.connect(leads_db_path)
    conn.row_factory = sqlite3.Row
    rows = conn.execute("SELECT id, company_name, domain, platform, source FROM leads").fetchall()
    leads = [dict(r) for r in rows]
    conn.close()

    # Load 9at contacts
    print(f"Loading 9at contacts from {nine_at_path}...")
    contacts_by_norm = load_9at_contacts(nine_at_path)
    print(f"Loaded {sum(len(v) for v in contacts_by_norm.values())} contacts across {len(contacts_by_norm)} companies")

    # Get all unique lead company names
    lead_names = list({l["company_name"] for l in leads})
    nine_at_names_norm = list(contacts_by_norm.keys())

    # Build a quick map from normalized lead name to originals
    print(f"Matching {len(lead_names)} unique lead names against {len(nine_at_names_norm)} 9at companies...")

    # Normalize lead names
    lead_norm_map = {name: normalize_company_name(name) for name in lead_names}

    # For each lead, try to find a match in the contacts dict
    results = []
    matched_leads = set()

    for lead in leads:
        lead_name = lead["company_name"]
        lead_norm = lead_norm_map[lead_name]

        if not lead_norm or len(lead_norm) < 2:
            continue

        lead_tokens = _extract_tokens(lead_norm)
        found_contacts = []
        match_key = None

        # Strategy 1: Exact normalized match
        if lead_norm in contacts_by_norm:
            found_contacts = contacts_by_norm[lead_norm]
            match_key = lead_norm
        else:
            # Strategy 2: Fuzzy matching using _is_good_match
            for nine_norm, contact_list in contacts_by_norm.items():
                if not nine_norm:
                    continue
                if _is_good_match(lead_norm, nine_norm):
                    found_contacts = contact_list
                    match_key = nine_norm
                    break

            # Strategy 3: Try variant splits for concatenated names (e.g. lcatterton)
            if not found_contacts:
                variants = _generate_variants(lead_name)
                for variant in variants[1:]:  # skip first, it's the same as lead_norm
                    if variant in contacts_by_norm:
                        found_contacts = contacts_by_norm[variant]
                        match_key = variant
                        break
                    for nine_norm, contact_list in contacts_by_norm.items():
                        if not nine_norm:
                            continue
                        if _is_good_match(variant, nine_norm):
                            found_contacts = contact_list
                            match_key = nine_norm
                            break
                    if found_contacts:
                        break

        if found_contacts:
            matched_leads.add(lead_name)
            results.append({
                "lead_id": lead["id"],
                "lead_company": lead_name,
                "lead_domain": lead["domain"],
                "lead_platform": lead["platform"],
                "lead_source": lead["source"],
                "match_key": match_key,
                "contacts": found_contacts,
            })

    print(f"Matched {len(matched_leads)} unique leads, found {sum(len(r['contacts']) for r in results)} contact records")
    return results


def save_contacts_to_db(
    results: list[dict],
    db_path: str = DB_PATH,
):
    """Save matched contacts to the contacts table in the database."""
    from db import init_contacts_table

    init_contacts_table()

    conn = sqlite3.connect(db_path)
    conn.execute("PRAGMA journal_mode=WAL")

    # Clear existing contacts (fresh match run)
    conn.execute("DELETE FROM contacts")

    count = 0
    for result in results:
        lead_id = result["lead_id"]
        for contact in result["contacts"]:
            conn.execute(
                """INSERT INTO contacts
                   (lead_id, first_name, last_name, title, email, email_status,
                    phone, linkedin, company_name_9at)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                (
                    lead_id,
                    contact["first_name"],
                    contact["last_name"],
                    contact["title"],
                    contact["email"],
                    contact["email_status"],
                    contact["phone"],
                    contact["linkedin"],
                    contact["company_name"],
                ),
            )
            count += 1

    conn.commit()
    conn.close()
    print(f"Saved {count} contacts to database")
    return count


if __name__ == "__main__":
    results = match_leads_to_contacts()

    print("\n--- Match Summary ---")
    matched_companies = set()
    total_contacts = 0
    for r in results:
        matched_companies.add(r["lead_company"])
        total_contacts += len(r["contacts"])

    print(f"Unique leads matched: {len(matched_companies)}")
    print(f"Total contact records: {total_contacts}")

    print("\n--- Sample Matches ---")
    for r in results[:15]:
        company = r["lead_company"]
        n_contacts = len(r["contacts"])
        sample = r["contacts"][0]
        nine_at_co = sample["company_name"]
        print(f"  {company} -> {nine_at_co} ({n_contacts} contacts)")
        for c in r["contacts"][:2]:
            print(f"    - {c['first_name']} {c['last_name']}, {c['title']}, {c['email']}")

    # Save to DB
    print("\nSaving contacts to database...")
    save_contacts_to_db(results)
