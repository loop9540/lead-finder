"""
Export leads to Excel with formatting.
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


HEADERS = [
    "Company Name",
    "Domain",
    "Competitor Platform",
    "Source",
]

# Will be extended when 9at contacts are integrated
CONTACT_HEADERS = [
    "Contact Name",
    "Contact Email",
    "Contact Title",
]


def export_to_excel(leads: list[dict], filepath: str, include_contacts: bool = False):
    """Export leads to a formatted Excel file."""
    os.makedirs(os.path.dirname(filepath), exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"

    headers = HEADERS + (CONTACT_HEADERS if include_contacts else [])

    # Style header row
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # Write data
    for row_idx, lead in enumerate(leads, 2):
        values = [
            lead.get("company_name", ""),
            lead.get("domain", ""),
            lead.get("platform", ""),
            lead.get("source", ""),
        ]
        if include_contacts:
            values += [
                lead.get("contact_name", ""),
                lead.get("contact_email", ""),
                lead.get("contact_title", ""),
            ]

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = thin_border

    # Auto-width columns
    for col in range(1, len(headers) + 1):
        max_len = max(
            len(str(ws.cell(row=r, column=col).value or ""))
            for r in range(1, len(leads) + 2)
        )
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = min(max_len + 4, 50)

    # Freeze header row
    ws.freeze_panes = "A2"

    # Add auto-filter
    ws.auto_filter.ref = ws.dimensions

    wb.save(filepath)
    print(f"  [export] Saved {len(leads)} leads to {filepath}")
